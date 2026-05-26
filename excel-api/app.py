from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)

def build_excel(r):
    wb = Workbook()
    ws = wb.active
    ws.title = "MESOCICLO"
    MICROS = 6  # MC1-MC6 normales + DESCARGA separada

    C_GREEN_DARK="00543A"; C_GREEN_MID="007A52"; C_GREEN_LIGHT="E8F5F0"
    C_WHITE="FFFFFF"; C_GRAY_MED="DDDDDD"; C_ORANGE="C0392B"
    C_ORANGE_LIGHT="FDECEA"; C_ORANGE_DARK="922B21"
    C_DARK="1A1A2E"; C_BLUE_LIGHT="EBF3FB"; C_BLUE_MED="D6E8F7"
    C_EJ_BG="F8F8F8"; C_SERIE_EVEN="FFFFFF"; C_SERIE_ODD="F2F6FF"

    def fill(h): return PatternFill("solid", fgColor=h)
    thin  = Side(style='thin',   color=C_GRAY_MED)
    med_g = Side(style='medium', color=C_GREEN_DARK)
    med_o = Side(style='medium', color=C_ORANGE)
    med_gr= Side(style='medium', color="999999")
    none  = Side(style=None)

    # COLUMNS:
    # A = ejercicio
    # B = series/rir normales
    # C..H = MC1..MC6 kg/reps
    # I = DESCARGA series/rir (nueva columna)
    # J = DESCARGA kg/reps
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 17

    col = 3
    mc_cols = []  # (col, num) for normal microciclos
    for m in range(MICROS):
        ws.column_dimensions[get_column_letter(col)].width = 15
        mc_cols.append((col, m+1))
        col += 1

    # Descarga: two cols (series/rir + kg/reps)
    desc_serie_col = col
    desc_kg_col    = col + 1
    ws.column_dimensions[get_column_letter(desc_serie_col)].width = 17
    ws.column_dimensions[get_column_letter(desc_kg_col)].width    = 15
    total_cols = desc_kg_col

    row = 1

    # ── TITLE ──
    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value="MESOCICLO")
    c.font = Font(bold=True, color=C_WHITE, size=15, name="Calibri")
    c.fill = fill(C_GREEN_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(left=med_g, top=med_g, bottom=med_g)
    for ci in range(2, total_cols+1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = fill(C_GREEN_DARK)
        cell.border = Border(top=med_g, bottom=med_g,
                             right=med_g if ci==total_cols else none)
    cc = max(total_cols-1, 3)
    ws.cell(row=row, column=cc).value = f"Cliente: {r.get('cliente','')}"
    ws.cell(row=row, column=cc).font = Font(bold=True, color=C_WHITE, size=10, name="Calibri")
    ws.cell(row=row, column=cc).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=row, start_column=cc, end_row=row, end_column=total_cols)
    row += 1

    # ── HEADER ROW ──
    ws.row_dimensions[row].height = 24
    for ci, label in [(1,"EJERCICIO"), (2,"SERIES  /  RIR")]:
        c = ws.cell(row=row, column=ci, value=label)
        c.font = Font(bold=True, color=C_WHITE, size=8.5, name="Calibri")
        c.fill = fill(C_GREEN_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=med_g if ci==1 else med_gr, bottom=med_g, right=med_gr)

    # Normal MC headers
    for mc, num in mc_cols:
        c = ws.cell(row=row, column=mc, value=f"MICROCICLO {num}")
        c.font = Font(bold=True, color=C_WHITE, size=8.5, name="Calibri")
        c.fill = fill(C_GREEN_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=med_gr, bottom=med_g, right=med_gr)

    # Descarga headers — two cells, merged label + kg/reps
    cd = ws.cell(row=row, column=desc_serie_col, value="DESCARGA\nSERIES  /  RIR")
    cd.font = Font(bold=True, color=C_WHITE, size=8.5, name="Calibri")
    cd.fill = fill(C_ORANGE)
    cd.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cd.border = Border(left=Side(style='medium', color=C_ORANGE_DARK),
                       bottom=med_o, top=med_o)

    ck = ws.cell(row=row, column=desc_kg_col, value="DESCARGA\nKG / REPS")
    ck.font = Font(bold=True, color=C_WHITE, size=8.5, name="Calibri")
    ck.fill = fill(C_ORANGE)
    ck.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ck.border = Border(bottom=med_o, top=med_o,
                       right=Side(style='medium', color=C_ORANGE_DARK))
    row += 1

    # ── DAYS ──
    for dia in r["dias"]:
        ws.row_dimensions[row].height = 20
        c = ws.cell(row=row, column=1,
                    value=f"  ENTRENAMIENTO DÍA {dia['numero']}  —  {dia['nombre']}")
        c.font = Font(bold=True, color=C_WHITE, size=9, name="Calibri")
        c.fill = fill(C_GREEN_MID)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(left=med_g, top=med_g, bottom=med_g)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

        for mc, num in mc_cols:
            cell = ws.cell(row=row, column=mc, value="KG / REPS")
            cell.font = Font(bold=True, color=C_WHITE, size=8, name="Calibri")
            cell.fill = fill(C_GREEN_MID)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=med_gr, top=med_g, bottom=med_g, right=med_gr)

        # Descarga day subheaders
        for ci in [desc_serie_col, desc_kg_col]:
            label = "SERIES / RIR" if ci==desc_serie_col else "KG / REPS"
            cell = ws.cell(row=row, column=ci, value=label)
            cell.font = Font(bold=True, color=C_WHITE, size=8, name="Calibri")
            cell.fill = fill(C_ORANGE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=med_gr if ci==desc_serie_col else none,
                                 top=med_o, bottom=med_o,
                                 right=Side(style='medium', color=C_ORANGE_DARK) if ci==desc_kg_col else none)
        row += 1

        # Exercises
        for ej in dia["ejercicios"]:
            sl = ej.get("series", [])
            if not isinstance(sl, list):
                sl = [{"tipo":"","series":str(sl),"reps":ej.get("reps",""),"rir":ej.get("rir","0")}]
            ns = len(sl)

            # Descarga series — increase RIR by 1, use high rep range
            desc_sl = ej.get("descarga_series", [])
            if not desc_sl:
                desc_sl = []
                for s in sl:
                    rir_d = min(int(s.get("rir","0"))+2, 5) if str(s.get("rir","0")).isdigit() else 2
                    reps_d = s.get("reps","")
                    num_d  = s.get("series","")
                    desc_sl.append({"series": num_d, "reps": reps_d, "rir": str(rir_d)})

            for si, s in enumerate(sl):
                ws.row_dimensions[row].height = 17

                # Col A merged
                if si == 0:
                    c = ws.cell(row=row, column=1, value=ej["nombre"])
                    c.font = Font(bold=True, color=C_DARK, size=8.5, name="Calibri")
                    c.fill = fill(C_EJ_BG)
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    c.border = Border(left=med_g, top=thin, bottom=thin, right=med_gr)
                    if ns > 1:
                        ws.merge_cells(start_row=row, start_column=1,
                                       end_row=row+ns-1, end_column=1)

                # Col B — normal series/rir
                reps  = s.get("reps","")
                rir   = s.get("rir","0")
                num_s = s.get("series","")
                bg_s  = C_SERIE_EVEN if si%2==0 else C_SERIE_ODD

                cb = ws.cell(row=row, column=2, value=f"{num_s}x{reps}  /  Rir{rir}")
                cb.font = Font(color="333333", size=8.5, name="Calibri")
                cb.fill = fill(bg_s)
                cb.alignment = Alignment(horizontal="center", vertical="center")
                cb.border = Border(left=med_gr, top=thin, bottom=thin, right=med_gr)

                # Normal MC cols
                rir_i = int(rir) if str(rir).isdigit() else 0
                for i, (mc, num) in enumerate(mc_cols):
                    if i >= MICROS-2:
                        bg_m = C_BLUE_MED
                    else:
                        bg_m = C_BLUE_LIGHT if i%2==0 else C_WHITE
                    cell = ws.cell(row=row, column=mc, value="")
                    cell.fill = fill(bg_m)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=med_gr, top=thin, bottom=thin, right=med_gr)

                # Descarga cols
                d = desc_sl[si] if si < len(desc_sl) else {"series":num_s,"reps":reps,"rir":"3"}
                d_label = f"{d.get('series','')}x{d.get('reps','')}  /  Rir{d.get('rir','3')}"

                cd = ws.cell(row=row, column=desc_serie_col, value=d_label)
                cd.font = Font(color=C_ORANGE, size=8.5, name="Calibri", bold=True)
                cd.fill = fill(C_ORANGE_LIGHT)
                cd.alignment = Alignment(horizontal="center", vertical="center")
                cd.border = Border(left=Side(style='medium', color=C_ORANGE_DARK),
                                   top=thin, bottom=thin)

                ck = ws.cell(row=row, column=desc_kg_col, value="")
                ck.fill = fill("FFF0EE")
                ck.border = Border(top=thin, bottom=thin,
                                   right=Side(style='medium', color=C_ORANGE_DARK))
                row += 1

            ws.row_dimensions[row].height = 3
            for ci in range(1, total_cols+1):
                ws.cell(row=row, column=ci).fill = fill("EEEEEE")
            row += 1

        ws.row_dimensions[row].height = 6
        for ci in range(1, total_cols+1):
            ws.cell(row=row, column=ci).fill = fill("CCCCCC")
        row += 1

    # ── COACHING NOTES ──
    if r.get("notas_coaching"):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value="  NOTAS DE COACHING")
        c.font = Font(bold=True, color=C_WHITE, size=9, name="Calibri")
        c.fill = fill(C_GREEN_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(left=med_g, top=med_g, bottom=med_g)
        for ci in range(2, total_cols+1):
            cell = ws.cell(row=row, column=ci)
            cell.fill = fill(C_GREEN_DARK)
            cell.border = Border(top=med_g, bottom=med_g,
                                 right=med_g if ci==total_cols else none)
        row += 1
        ws.row_dimensions[row].height = 55
        c = ws.cell(row=row, column=1, value=r["notas_coaching"])
        c.font = Font(color=C_DARK, size=8.5, name="Calibri", italic=True)
        c.fill = fill(C_GREEN_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        c.border = Border(left=med_g, bottom=med_g, right=med_g)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)

    ws.freeze_panes = "C3"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/', methods=['GET'])
def health():
    return jsonify({"status": "ok", "service": "Flowix Excel Generator"})


@app.route('/excel', methods=['POST', 'OPTIONS'])
def generar_excel():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        data = request.get_json()
        if not data or 'dias' not in data:
            return jsonify({"error": "JSON inválido"}), 400
        excel_file = build_excel(data)
        nombre = (data.get('cliente') or 'cliente').replace(' ', '_')
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Rutina_{nombre}_Flowix.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
